import openpyxl
#pop install openpyxl #bibliotek do excela

workbook = openpyxl.load_workbook('sales.xlsx') #wczytywanie z pliku do zmiennej workbook
#workbook = openpyxl.load_workbook('D// new folder// : sales.xlsx') #wczytywanie z pliku do zmiennej workbook
worksheet = workbook.active # otwiera ostatni uzywany arkusz
#worksheet = workbook.active ('Arkusz 3')#

print(worksheet) #worksheet "arkusz1

lista = []

for w in worksheet:  #obiekty
    print(w)

for i in range(0, worksheet.max_row):       #
        for col in worksheet.iter_cols(1, worksheet.max_column):
            lista.append(col[i].value)
print(lista) #cala lista
print(lista[0:3]) #pierwsze trzy elelmennty z pierwszego wiersza

for i in range(0, len(lista), 3): #petla idzie co trzy gdyz sa po 3 ellementy w linii
    print(lista[i:i+3])
